from openpyxl import load_workbook
from rows_count import row_count

# The spreadsheet used in the  operator and manager functions is only defined in the spreadsheet function
# It is the called into other functions

#Spreadsheet named
def spreadsheet():
    file = 'xlwfile.xlsx' # Only added to this function
    return file


# Sheet 1 of the spreadsheet
def operator():
    file = spreadsheet() 
    r = row_count(file, 'B') # Calling the count function on the file named, the 'B' just chooses a column
    wb = load_workbook(filename = file) # wb is loading the workbook as a var
    sheet_ranges = wb['operators'] # This is accessing the sheet of the workbook
    c =[]

    for oisam in range(2,r):
        a = 'B'+ str(oisam)
        b = [(sheet_ranges[a].value)]
        c = c + b                       # Changes the column into as array
    print("The operator ISAMs are ", c)  
    print()


# Sheet two of the spreadsheet    
def manager():
    file = spreadsheet()
    r = row_count(file,'B') # Calling the count function on the file called
    wb = load_workbook(filename = file) # wb is loading the workbook as a var
    sheet_ranges = wb['act'] # This is accessing the sheet of the workbook
    f =[]

    for misam in range(2,r):
        d = 'B'+ str(misam)
        e = [(sheet_ranges[d].value)]
        f = f + e                       # Changes the column into as array
    print("The manager ISAMs are ", f)
    


if __name__ == '__main__':
    operator()
    manager()
